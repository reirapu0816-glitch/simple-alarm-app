import csv, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "座席表.csv")
OUT = os.path.join(ROOT, "座席表.xlsx")

with open(SRC, encoding="utf-8") as f:
    rows = [r for r in csv.reader(f) if r and r[0].strip()]
header, data = rows[0], rows[1:]

FONT = "Meiryo"
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1F4E79")
head_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
body_font = Font(name=FONT, size=11)
alt_fill = PatternFill("solid", fgColor="F2F7FB")

wb = Workbook()

# ---------- Sheet 1: 座席リスト ----------
ws = wb.active
ws.title = "座席リスト"
ws.append(header)
for c in ws[1]:
    c.font, c.fill, c.border = head_font, head_fill, border
    c.alignment = Alignment(horizontal="center", vertical="center")

for i, r in enumerate(data, start=2):
    ws.append(r)
    for j in range(1, 4):
        c = ws.cell(row=i, column=j)
        c.font, c.border = body_font, border
        c.alignment = Alignment(horizontal="center" if j != 2 else "left", vertical="center")
        if i % 2 == 0:
            c.fill = alt_fill

last = len(data) + 1
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{last}"
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 14
ws.row_dimensions[1].height = 24

# ---------- Sheet 2: 座席表(配置図) ----------
def key(s):
    m = re.match(r"([A-Z]+)-(\d+)$", s)
    return m.group(1), int(m.group(2))

blocks, maxcol = [], 0
for r in data:
    b, n = key(r[0])
    if b not in blocks:
        blocks.append(b)
    maxcol = max(maxcol, n)

lookup = {r[0]: (r[1], r[2]) for r in data}
g = wb.create_sheet("座席表")
g["A1"] = "座席配置図（各セル：氏名／期）"
g["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
g.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxcol + 1)

HR = 3  # header row of the grid
g.cell(row=HR, column=1, value="列＼番号")
for n in range(1, maxcol + 1):
    g.cell(row=HR, column=n + 1, value=n)
for c in g[HR][: maxcol + 1]:
    c.font, c.fill, c.border = head_font, head_fill, border
    c.alignment = Alignment(horizontal="center", vertical="center")

for bi, b in enumerate(blocks):
    row = HR + 1 + bi
    hc = g.cell(row=row, column=1, value=b)
    hc.font, hc.fill, hc.border = head_font, head_fill, border
    hc.alignment = Alignment(horizontal="center", vertical="center")
    for n in range(1, maxcol + 1):
        seat = f"{b}-{n}"
        cell = g.cell(row=row, column=n + 1)
        rec = lookup.get(seat)
        cell.value = f"{rec[0]}\n{rec[1]}" if rec else None
        cell.font = Font(name=FONT, size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

g.column_dimensions["A"].width = 10
for n in range(1, maxcol + 1):
    g.column_dimensions[get_column_letter(n + 1)].width = 16
for bi in range(len(blocks)):
    g.row_dimensions[HR + 1 + bi].height = 34
g.freeze_panes = g.cell(row=HR + 1, column=2).coordinate

note = HR + len(blocks) + 2
g.cell(row=note, column=1, value="※ この配置図は「座席リスト」シートの内容をもとに作成しています。空白セルは該当座席なしです。").font = Font(name=FONT, size=9, italic=True, color="808080")

# ---------- Sheet 3: 期別集計 ----------
def sortkey(v):
    m = re.match(r"(\d+)期", v)
    return (0, int(m.group(1)), v) if m else (1, 0, v)

counts = Counter(r[2] for r in data)
kinds = sorted(counts, key=sortkey)
s = wb.create_sheet("期別集計")
s["A1"] = "何期生"
s["B1"] = "人数"
for c in (s["A1"], s["B1"]):
    c.font, c.fill, c.border = head_font, head_fill, border
    c.alignment = Alignment(horizontal="center", vertical="center")
for i, k in enumerate(kinds, start=2):
    s.cell(row=i, column=1, value=k)
    s.cell(row=i, column=2, value=counts[k])
    for j in (1, 2):
        c = s.cell(row=i, column=j)
        c.font, c.border = body_font, border
        c.alignment = Alignment(horizontal="center" if j == 2 else "left", vertical="center")
tot = len(kinds) + 2
s.cell(row=tot, column=1, value="合計")
s.cell(row=tot, column=2, value=len(data))
for j in (1, 2):
    c = s.cell(row=tot, column=j)
    c.font = Font(name=FONT, bold=True)
    c.border = border
    c.alignment = Alignment(horizontal="center" if j == 2 else "left", vertical="center")
s.column_dimensions["A"].width = 16
s.column_dimensions["B"].width = 10
s.freeze_panes = "A2"

wb.save(OUT)
print("rows:", len(data), "blocks:", blocks, "maxcol:", maxcol, "kinds:", len(kinds))
print(Counter(r[2] for r in data).most_common(5))
